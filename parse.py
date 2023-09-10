from collections import Generator
from itertools import product
from typing import Union, List

import pandas as pd
import datetime

import openpyxl as xl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.styles.colors import Color
from openpyxl.worksheet.merge import MergedCellRange

pd.set_option('max_rows', None)  # print row without collapsing
pd.set_option('display.width', None)  # print columns without collapsing
pd.set_option('max_colwidth', None)  # print every cells without collapsing


class ParseObject:
    def __init__(self):
        self.students = dict()
        self.employee = dict()
        self.additional = dict()

    def to_dict(self):
        return {
            'students': self.students,
            'employee': self.employee,
            'additional': self.additional
        }

    def __str__(self):
        return f"ParseObject\n" \
               f"students: {self.students}\n" \
               f"employee: {self.employee}\n" \
               f"additional: {self.additional}\n"


class ExcelParser:
    # @staticmethod
    # def parse(file_path) -> ParseObject:
    #     excel: pd.DataFrame = pd.read_excel(file_path, engine='openpyxl')
    #
    #     print(len(excel.columns))
    #
    #     menus_with_target: pd.DataFrame = excel.iloc[3:11]  # 대상 포함
    #     menus_without_target: pd.DataFrame = menus_with_target.iloc[:, 1:]  # 대상 미포함
    #
    #     menus: pd.DataFrame = menus_without_target.transpose()
    #     menus.drop(4, axis=1, inplace=True)  # drop useless column
    #
    #     res = ParseObject()
    #     cur_year = datetime.datetime.today().year
    #
    #     for daily_menus in menus.values:
    #         daily_menus = list(daily_menus)
    #         # format
    #         # [date, ...menus, std/emp_only, additional]
    #
    #         date = daily_menus[0]  # dates
    #
    #         base_menus = daily_menus[1:-1]
    #
    #         base_for_student = base_menus[:-1]  # student menu
    #         base_for_student.append(base_menus[-1].split('/')[0])
    #
    #         base_for_employee = base_menus[:-1]  # employee menu
    #         base_for_employee.extend(base_menus[-1].split('/'))
    #
    #         # additional_menus = daily_menus[-1].split('\n')  # additional menu
    #
    #         key = f'{cur_year}-{date[4:9].replace(".", "-")}'
    #
    #         res.students[key] = base_for_student
    #         res.employee[key] = base_for_employee
    #         # res.additional[key] = additional_menus
    #     print(res)
    #     return res

    @staticmethod
    def _get_source_of_merged_cells(sheet: Worksheet, cell: MergedCell) -> Cell:
        merged_ranges: set = sheet.merged_cells.ranges
        merged_range: MergedCellRange
        for merged_range in merged_ranges:  # each merge ranges
            cell_range = list(merged_range.cells)
            base_cell = cell_range[0]  # get base cell among the merged cells
            cell_coord = (cell.row, cell.column)
            if cell_coord in cell_range:  # check if target cell is contained in this range
                return sheet.cell(row=base_cell[0], column=base_cell[1])  # return base_cell's value
        raise Exception(f"Failed to find Base Cell of {cell}")

    @staticmethod
    def _get_cell_source(sheet: Worksheet, cell: Union[Cell, MergedCell]) -> Cell:
        if type(cell) == MergedCell:
            cell: Cell = ExcelParser._get_source_of_merged_cells(sheet, cell)
        return cell


    @staticmethod
    def parse(file_path: str) -> ParseObject:
        res = ParseObject()

        excel_wb: Workbook = xl.load_workbook(file_path, data_only=True, rich_text=True)
        if len(excel_wb.sheetnames) == 1:
            excel_ws: Worksheet = excel_wb[excel_wb.sheetnames[0]]
        # have to handle multiple worksheets if exists.

        columns: list = [column[4:] for column in list(excel_ws.columns)]
        indices: list = columns[:1]
        datas: list = columns[1:]

        # get employee color
        employee_color: str = ''
        cell: [Cell, MergedCell]
        for cell in indices[0]:
            if cell.value is None:
                continue
            if type(cell.value) == CellRichText:
                value: CellRichText = cell.value
                item: TextBlock
                for item in value:
                    if item.text == '교직원':
                        employee_color = item.font.color.rgb

        # parse data
        column: list
        for column in datas:
            cell: Union[Cell, MergedCell]
            value_list: list = []
            for cell in column:
                if cell.value is None:
                    continue
                if type(cell.value) == CellRichText:
                    values: CellRichText = cell.value
                    for text in values:
                        if type(text) is not TextBlock:  # default text
                            continue
                        if text.font.color is None:  # default text
                            continue
                        color: Color = text.font.color
                        print(color.rgb)
                        print(text.text)
                else:
                    value: str = cell.value
                value_list.append(value)
            # print(value_list)

        return res

    # def test_templates(self):
    #     cur_excel: pd.DataFrame = pd.read_excel('./datas/20230902.xlsx', engine='openpyxl')
    #     test1_excel: pd.DataFrame = pd.read_excel('./datas/test1.xlsx', engine='openpyxl')
    #     test2_excel: pd.DataFrame = pd.read_excel('./datas/test2.xlsx', engine='openpyxl')
    #
    #     cur_excel.dropna(how='all', axis=1, inplace=True)
    #     test1_excel.dropna(how='all', axis=1, inplace=True)
    #     test2_excel.dropna(how='all', axis=1, inplace=True)
    #
    #     cur_excel.dropna(how='all', inplace=True)
    #     test1_excel.dropna(how='all', inplace=True)
    #     test2_excel.dropna(how='all', inplace=True)
    #
    #     print(len(cur_excel.columns))
    #     print(len(test1_excel.columns))
    #     print(len(test2_excel.columns))
    #
    #     print(cur_excel.shape[0])
    #     print(test1_excel.shape[0])
    #     print(test2_excel.shape[0])
    #
    #     cur_excel = cur_excel.transpose()
    #     test1_excel = test1_excel.transpose()
    #     test2_excel = test2_excel.transpose()
    #
    #     print(cur_excel.values)
    #     print(test1_excel.values)
    #     print(test2_excel.values)


if __name__ == '__main__':
    parser = ExcelParser()
    # parser.parse('./datas/20230902.xlsx')
    # parser.parse('./datas/test1.xlsx')
    # parser.parse('./datas/test2.xlsx')
    # parser.test_templates()
    parser.parse('./datas/0911-0915.xlsx')
